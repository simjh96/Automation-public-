import time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from soynlp import hangle
import re
import os


class ToExcel():
    """
    to_excel(보통주기준)
        dart
        kisline
            img_loader
        *chrome
            업무번호
            업체정보
                과거 내역
        *sql
            daily 여부

    - 자료 -> out_dir(최종으로 굽는곳) -> 개별적으로 검수 후 가져가

    - init에 action 없이 함수만 빼서 사용하는게 좋을듯

    """

    #0계층
    def __init__(self, upmoo_name_clean, out_dir):
        #variables
        self.upmoo_name_clean = upmoo_name_clean      
        self.out_dir = out_dir

        self.auto_xl()
    
    #1계층
    def auto_xl(self):
        """
        가장 먼저 실행해야하는 파일 생성 기능
            AUTO sheet 생성
            out_dir + \\(KISP){}_AUTO.xlsx 생성
        """
        os.makedirs(self.out_dir,exist_ok=True)
        wb = Workbook()
        ws = wb.create_sheet("AUTO")
        try:
            wb.save(self.out_dir + "\\(KISP){}_AUTO.xlsx".format(self.upmoo_name_clean))
        except:
            print('시스템>>>{} 관련 엑셀이 이미 존재합니다!'.format(self.upmoo_name_clean))

    
    def dfs_to_sheet(self, dfs_dic,sheet_name,hyper=True):
        """
        dfs_dic = {delimiter:df,delimiter:df,...}

        dfs_dic -> add to sheet -> save
        handling hyper link option needs to be developed...
        """
        #load file
        try:
            wb = load_workbook(self.out_dir + "\\(KISP){}_AUTO.xlsx".format(self.upmoo_name_clean))
        except:
            print('시스템>>>{} 관련 엑셀이 존재하지 않습니다!'.format(self.upmoo_name_clean))
            return None
        ws = wb.create_sheet(sheet_name)

        #loop through dfs_dic
        delimiters = list(dfs_dic.keys())
    
        for delimiter in delimiters:
            ws.append(['# '+delimiter])
            ws.append(['======','======','======','======','======','======','======'])

            #자료 없으면 pass(빈 df는 알아서 0줄))
            if dfs_dic[delimiter] is None:
                pass
            else:
                for r in dataframe_to_rows(dfs_dic[delimiter], index=True, header=True):
                    ws.append(r)
            ws.append(['======','======','======','======','======','======','======'])
                
        wb.save(self.out_dir + "\\(KISP){}_AUTO.xlsx".format(self.upmoo_name_clean))
        print('시스템>>>{}의 {} 시트(df) 생성 후 저장 완료'.format(self.upmoo_name_clean, sheet_name))
        time.sleep(2)
        wb.close()


    def ss_to_sheet(self,second_name,sheet_name,memo):
        """
        upmoo_name_clean -> search ss -> add to sheet -> save
        memo added
            ex) 신용등급 BB+

        겹쳐 있으니 하나씩 옮겨야함
        """
        os.makedirs(self.out_dir + '\\screenshots',exist_ok=True)
        
        #load file
        try:
            wb = load_workbook(self.out_dir + "\\(KISP){}_AUTO.xlsx".format(self.upmoo_name_clean))
        except:
            print('시스템>>>{} 관련 엑셀이 존재하지 않습니다!'.format(self.upmoo_name_clean))
            return None
        ws = wb.create_sheet(sheet_name)

        #all screenshots in screenshots folder(need time to download)
        time.sleep(5)
        sss = []
        for file in os.listdir(self.out_dir + '\\screenshots'):
            if ((self.upmoo_name_clean +'_'+ second_name) in file):
                sss.append(file)

        print('시스템>>> {} 첨부 시작합니다'.format(str(sss)))    
        for ss in sss:
            img = openpyxl.drawing.image.Image(self.out_dir + '\\screenshots' + '\\' + ss)
            img.anchor = 'A1'
            ws.add_image(img)

        #single row can not be appended...
        #use second memo as previous rating?
        ws.append([memo])            
        wb.save(self.out_dir + "\\(KISP){}_AUTO.xlsx".format(self.upmoo_name_clean))
        print('시스템>>>{}의 {} 정보와 {} 시트(이미지) 생성 후 저장 완료'.format(self.upmoo_name_clean, second_name, sheet_name))
        time.sleep(2)
        wb.close()

# hyperlink 작업 자료
# #감사보고서 자료 넣는 과정
# dfs = search_audit_list(upchae_name)

# #url on
# dfs[0].loc[:,'links'] = dfs[0].loc[:,'links'].apply(lambda x: url_dic['dart_search'] + x)

# input_dfs = [dfs[0], dfs[1][0], dfs[1][1], dfs[1][2], dfs[1][3]]
# delimiters = ['page_list', 'most recent BS', 'IS', 'Capital Change', 'Shareholder List']

# df_to_adeq_sheet(delimiters,input_dfs,upchae_name,'DART')



# #search name_cells 
# idx_cell = None
# for row in ws.iter_rows():
#     for cell in row:
#         if cell.value == '종목명[노트용]':
#             idx_cell = cell
# for i in range(ws.max_row):
#     if idx_cell.offset(i+1,0).value != None:
#         work_sec_name_cells.append(idx_cell.offset(i+1,0))




