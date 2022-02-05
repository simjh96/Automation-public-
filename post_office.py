import time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import shutil
import os
import re


class PostOffice():
    """

    #어차피 중복 제거된 종목들의 검색 자료 토대로 생성할거임
    #즉, 1 -> n 대응 함수임, input 측면은 무조건 unique 값임

    ex)
    po0 = post_office.PostOffice('회사명1'
    ,'업무배분 (15).xlsx'
    , '심재훈'
    ,r'1. 보통주'
    ,r'\\평가내역')
    po0.copy_for_upmoo()
    po0.post_and_clear()


    """

    # 0계층
    def __init__(self, upmoo_name_clean, up_bae_total_dir, upmoo_damdang, posted_from, posted_to):
        # variables
        self.upmoo_name_clean = upmoo_name_clean
        self.up_bae_total_dir = up_bae_total_dir
        self.posted_from = posted_from
        self.posted_to = posted_to
        self.upmoo_damdang = upmoo_damdang

    # 1계층
    def copy_for_upmoo(self, sign=True):
        """
        upmoo_name_clean = result.loc[i,'종목명(업무배분_clean)']
        upchae_name은 파일 이름과도 동일
        up_bae_total_dir = 업무배분 (14) xlsx
        posted_from 안에서 복사본을 만들것
        """

        print('시스템>>> 동일 엑셀 파일을 업무배분 가짓수만큼 복수, 이름 변경 합니다')
        # 업체 이름 입력 받아서 업무배분표와 대조해

        up_bae_total = pd.read_excel(r'{}'.format(self.up_bae_total_dir))
        bool0 = up_bae_total.loc[:, '종목명'].apply(
            self.sc_type_strip) == self.upmoo_name_clean
        up_bae_target = up_bae_total.loc[bool0, :]

        # 해당 파일명을 '_종목명_유형' 붙여서 복사해
        for i in range(len(up_bae_target)):
            shutil.copy2(self.posted_from + r'\(KISP){}_AUTO.xlsx'.format(self.upmoo_name_clean), self.posted_from + r'\(KISP){}_{}_{}_AUTO.xlsx'.format(
                self.upmoo_name_clean, up_bae_target.loc[:, ['유형']].iloc[i, 0], up_bae_target.loc[:, ['업무번호']].iloc[i, 0]))

            if sign:
                self.sign_upmoo(self.posted_from + r'\(KISP){}_{}_{}_AUTO.xlsx'.format(self.upmoo_name_clean, up_bae_target.loc[:, [
                                '유형']].iloc[i, 0], up_bae_target.loc[:, ['업무번호']].iloc[i, 0]), up_bae_target.iloc[[i], :])

        # 원본 파일 지워(사진복사안돼서 지우면 안됨)
        # os.remove(self.posted_from + r'\(KISP){}_AUTO.xlsx'.fromat(self.upmoo_name_clean))

    def post_and_clear(self):
        """
        post(empty OK) & clear all contents in folder
        ex)
        posted_from = '\\fresh_baked'
        posted_to = '\\평가내역'
        upmoo_damdang = '심재훈'
        upmoo_name_clean = '회사명1'
        """
        print('시스템>>> 송부 위치의 내용을 전부 송신위치로 옮깁니다.')
        # post
        try:
            shutil.copytree(self.posted_from, self.posted_to +
                            '\\{}(제작중)\\{}_AUTO'.format(self.upmoo_damdang, self.upmoo_name_clean))
            print('시스템>>> "{}" 이동 완료'.format(self.upmoo_name_clean))
        except:
            # 이미 업체 이름의 폴더가 있는 경우
            print('시스템>>> "{}" 중복 종목입니다... 관련 업무는 모두 auto 엑셀의 sheet1에 기입되어 있습니다'.format(
                self.upmoo_name_clean))

        # clear total folder and mk empty folder
        shutil.rmtree(self.posted_from)
        os.makedirs(self.posted_from, exist_ok=True)

    # 2계층
    def sign_upmoo(self, xl_dir, row_content_df):
        print('시스템>>> 엑셀의 Sheet1 에 업무배분 내용을 기재합니다')
        wb = load_workbook(xl_dir)
        ws = wb.create_sheet('업무배분')
        for r in dataframe_to_rows(row_content_df, index=True, header=True):
            ws.append(r)

        wb.save(xl_dir)
        print('시스템>>>업무배분 기입 후 저장 완료')
        time.sleep(1)
        wb.close()

    # def basic_strip(self, with_zoo, space=True):
    #     result = re.sub(r'\(.+?\)','',with_zoo.replace('㈜','')).strip()
    #     if space:
    #         return result
    #     else:
    #         return result.replace('\s','')

    def sc_type_strip(self, name):
        """
        strip ['RCPS','BW','CB','CPS','보통주','우선주','평가제외','제외']
        + re.sub(r'\(.+?\)','',name.replace('㈜','')).strip()
        """

        for i in ['RCPS', 'BW', 'CB', 'CPS', '보통주', '우선주', '상환', '전환', '교환권', '교환', '교부', '채권', '평가제외', '제외']:
            name = name.replace(i, '')
            name = name.strip()

        name = re.sub(r'\(.+?\)', '', name.replace('㈜', '')).strip()
        return name
